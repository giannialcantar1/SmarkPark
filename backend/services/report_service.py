from __future__ import annotations

from collections import defaultdict
from datetime import datetime
import io
import json
from pathlib import Path
import zipfile

from openpyxl import Workbook

from repositories import ParkingSpaceRepository, PaymentRepository, SessionRepository, UserRepository, VehicleRepository
from utils.excel_report_template import render_report_sheet
from utils.supabase_client import normalize_text, parse_datetime, select_rows, utcnow


class ReportService:
    def __init__(self) -> None:
        self.space_repository = ParkingSpaceRepository()
        self.vehicle_repository = VehicleRepository()
        self.session_repository = SessionRepository()
        self.payment_repository = PaymentRepository()
        self.user_repository = UserRepository()

    def _spaces(self, *, garage_id: str) -> list[dict]:
        return self.space_repository.get_all(
            filters=[{"column": "garage_id", "value": garage_id, "optional": False}],
            order_candidates=["piso", "numero"],
        )

    def _vehicles(self, *, garage_id: str) -> list[dict]:
        return [self._with_vehicle_display(row) for row in self.vehicle_repository.get_by_garage(garage_id)]

    def _sessions(self, *, garage_id: str) -> list[dict]:
        spaces = {str(row.get("id")) for row in self._spaces(garage_id=garage_id)}
        vehicle_rows = self._vehicles(garage_id=garage_id)
        vehicles = {str(row.get("id")) for row in vehicle_rows}
        vehicles_by_id = {str(row.get("id")): row for row in vehicle_rows if row.get("id")}
        result: list[dict] = []
        for row in self.session_repository.get_all(order_candidates=["entrada", "created_at"], desc=True):
            if row.get("garage_id") == garage_id or str(row.get("space_id")) in spaces or str(row.get("vehicle_id")) in vehicles:
                vehicle = vehicles_by_id.get(str(row.get("vehicle_id") or ""))
                result.append(self._with_session_vehicle_display(row, vehicle))
        return result

    def generate_occupancy_report(self, *, garage_id: str) -> dict:
        spaces = self._spaces(garage_id=garage_id)
        total = len(spaces)
        occupied = sum(1 for row in spaces if row.get("occupied"))
        available = max(total - occupied, 0)
        floors: dict[str, dict[str, int]] = defaultdict(lambda: {"occupied": 0, "available": 0})

        for row in spaces:
            floor = row.get("floor") or "Sin piso"
            if row.get("occupied"):
                floors[floor]["occupied"] += 1
            else:
                floors[floor]["available"] += 1

        floor_stats = []
        for floor in sorted(floors):
            floor_total = floors[floor]["occupied"] + floors[floor]["available"]
            floor_stats.append(
                {
                    "floor": floor,
                    "occupied": floors[floor]["occupied"],
                    "available": floors[floor]["available"],
                    "percentage": round((floors[floor]["occupied"] / floor_total) * 100) if floor_total else 0,
                }
            )

        return {
            "total_spaces": total,
            "occupied_spaces": occupied,
            "available_spaces": available,
            "occupancy_percentage": round((occupied / total) * 100) if total else 0,
            "floor_stats": floor_stats,
        }

    def generate_income_report(self, *, garage_id: str) -> dict:
        session_ids = {str(row.get("id")) for row in self._sessions(garage_id=garage_id) if row.get("id")}
        payments = self.payment_repository.get_all(order_candidates=["fecha", "created_at"], desc=True)
        total = 0.0
        today_total = 0.0
        today = utcnow().date()

        for row in payments:
            session_id = str(row.get("session_id") or "")
            if session_id not in session_ids:
                continue
            amount = float(row.get("monto") or row.get("amount") or 0)
            total += amount
            paid_at = parse_datetime(row.get("fecha") or row.get("created_at"))
            if paid_at and paid_at.date() == today:
                today_total += amount

        return {
            "total_income": round(total, 2),
            "today_income": round(today_total, 2),
            "payments_count": len([row for row in payments if str(row.get("session_id") or "") in session_ids]),
        }

    def generate_vehicle_report(self, *, garage_id: str) -> dict:
        vehicles = self._vehicles(garage_id=garage_id)
        summary = {"total": len(vehicles), "by_type": {}, "items": vehicles}
        for row in vehicles:
            vehicle_type = row.get("type") or "Sin tipo"
            summary["by_type"][vehicle_type] = summary["by_type"].get(vehicle_type, 0) + 1
        return summary

    @staticmethod
    def _vehicle_name(row: dict) -> str:
        if not row:
            return ""
        combined = str(
            row.get("marca_modelo")
            or row.get("brand_model")
            or row.get("vehicle_name")
            or row.get("nombre_vehiculo")
            or ""
        ).strip()
        if combined:
            return combined
        brand = str(row.get("marca") or row.get("brand") or "").strip()
        model = str(row.get("modelo") or row.get("model") or "").strip()
        return " ".join(part for part in (brand, model) if part).strip()

    @classmethod
    def _vehicle_display(cls, row: dict) -> str:
        if not row:
            return ""
        plate = str(row.get("placa") or row.get("plate") or "").strip().upper()
        vehicle_name = cls._vehicle_name(row)
        if plate and vehicle_name:
            return f"{plate} - {vehicle_name}"
        return plate or vehicle_name

    @classmethod
    def _with_vehicle_display(cls, row: dict) -> dict:
        enriched = dict(row)
        vehicle_name = cls._vehicle_name(enriched)
        vehicle_display = cls._vehicle_display(enriched)
        if vehicle_name:
            enriched.setdefault("vehiculo", vehicle_name)
            enriched.setdefault("vehicle_name", vehicle_name)
        if vehicle_display:
            enriched["vehiculo_display"] = vehicle_display
            enriched["vehicle_display"] = vehicle_display
        return enriched

    @classmethod
    def _with_session_vehicle_display(cls, row: dict, vehicle) -> dict:
        enriched = dict(row)
        if vehicle:
            enriched.setdefault("placa", vehicle.get("placa") or vehicle.get("plate"))
            enriched.setdefault("plate", vehicle.get("plate") or vehicle.get("placa"))
            enriched.setdefault("marca", vehicle.get("marca") or vehicle.get("brand"))
            enriched.setdefault("brand", vehicle.get("brand") or vehicle.get("marca"))
            enriched.setdefault("modelo", vehicle.get("modelo") or vehicle.get("model"))
            enriched.setdefault("model", vehicle.get("model") or vehicle.get("modelo"))
        display_source = {**(vehicle or {}), **enriched}
        vehicle_name = cls._vehicle_name(display_source)
        vehicle_display = cls._vehicle_display(display_source)
        if vehicle_name:
            enriched.setdefault("vehiculo", vehicle_name)
            enriched.setdefault("vehicle_name", vehicle_name)
        if vehicle_display:
            enriched["vehiculo_display"] = vehicle_display
            enriched["vehicle_display"] = vehicle_display
        return enriched

    def generate_user_report(self, *, garage_id: str) -> dict:
        users = self.user_repository.get_all(
            filters=[{"column": "garage_id", "value": garage_id, "optional": False}],
            order_candidates=["created_at", "nombre", "email"],
            desc=True,
        )
        summary = {"total": len(users), "by_role": {}, "items": users}
        for row in users:
            role = normalize_text(row.get("rol") or row.get("role")) or "usuario"
            summary["by_role"][role] = summary["by_role"].get(role, 0) + 1
        return summary

    def build_power_bi_import_bundle(
        self,
        *,
        payload: dict,
        garage_id: str,
        generated_by: dict | None = None,
    ) -> tuple[io.BytesIO, str]:
        tables = payload.get("tables") if isinstance(payload.get("tables"), dict) else {}
        table_rows = {
            "Resumen": self._coerce_rows(tables.get("Resumen") or payload.get("Resumen") or payload.get("resumen")),
            "Historico": self._coerce_rows(tables.get("Historico") or payload.get("Historico") or payload.get("historico")),
            "Ocupacion": self._coerce_rows(tables.get("Ocupacion") or payload.get("Ocupacion") or payload.get("ocupacion")),
            "Ingresos": self._coerce_rows(tables.get("Ingresos") or payload.get("Ingresos") or payload.get("ingresos")),
        }

        if not any(table_rows.values()):
            raise ValueError("No se recibieron tablas exportables para Power BI.")

        export_id = str(payload.get("export_id") or utcnow().strftime("%Y%m%d%H%M%S"))
        exported_at = str(payload.get("exported_at") or utcnow().isoformat())
        period_label = str(payload.get("period_label") or self._first_value(table_rows["Resumen"], "period_label") or "Periodo")
        floor_label = str(payload.get("floor_label") or self._first_value(table_rows["Resumen"], "floor_label") or "todos-los-pisos")
        relationships = payload.get("relationships") if isinstance(payload.get("relationships"), list) else []

        workbook_bytes = self._build_power_bi_workbook(
            table_rows=table_rows,
            relationships=relationships,
            export_id=export_id,
            exported_at=exported_at,
            garage_id=garage_id,
            generated_by=generated_by or {},
        )

        bundle_payload = {
            "version": payload.get("version") or "1.0",
            "source": payload.get("source") or "SmartPark Reports",
            "export_id": export_id,
            "exported_at": exported_at,
            "garage_id": garage_id,
            "format": "power-bi-import-bundle",
            "requested_format": "pbix",
            "tables": table_rows,
            "relationships": relationships,
        }

        safe_period = self._slugify(period_label)
        safe_floor = self._slugify(floor_label)
        zip_filename = f"smartpark-powerbi-bundle-{safe_period}-{safe_floor}.zip"

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("README.txt", self._build_power_bi_readme(period_label=period_label, floor_label=floor_label))
            archive.writestr("manifest.json", json.dumps(bundle_payload, ensure_ascii=False, indent=2).encode("utf-8"))
            archive.writestr("relationships.json", json.dumps(relationships, ensure_ascii=False, indent=2).encode("utf-8"))
            archive.writestr("smartpark-powerbi.json", json.dumps(bundle_payload, ensure_ascii=False, indent=2).encode("utf-8"))
            archive.writestr("smartpark-powerbi.xlsx", workbook_bytes.getvalue())

        zip_buffer.seek(0)
        return zip_buffer, zip_filename

    def build_parking_report_workbook(
        self,
        *,
        payload: dict,
        garage_id: str,
        generated_by: dict | None = None,
    ) -> tuple[io.BytesIO, str]:
        rows = self._coerce_report_rows(payload.get("rows"))
        if not rows:
            raise ValueError("No se recibieron filas exportables para el reporte.")

        title = str(payload.get("title") or "Reporte de Parkings")
        generated_at = parse_datetime(payload.get("generated_at")) or utcnow()
        garage = self._garage_info(garage_id)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Parkings"
        render_report_sheet(
            worksheet,
            title=title,
            rows=rows,
            garage=garage,
            generated_by=generated_by or {},
            generated_at=generated_at,
            table_name="SmartPark_Parkings",
            logo_path=self._default_logo_path(),
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"{self._slugify(title)}-{generated_at:%Y-%m-%d}.xlsx"
        return output, filename

    @staticmethod
    def _coerce_rows(value) -> list[dict]:
        if not isinstance(value, list):
            return []
        return [row for row in value if isinstance(row, dict)]

    @classmethod
    def _coerce_report_rows(cls, value) -> list[dict]:
        if not isinstance(value, list):
            return []
        rows: list[dict] = []
        for row in value:
            if isinstance(row, dict):
                rows.append(row)
                continue
            if isinstance(row, str):
                parsed = cls._parse_delimited_row(row)
                if parsed:
                    rows.append(parsed)
        return rows

    @staticmethod
    def _parse_delimited_row(value: str) -> dict:
        import csv

        text = str(value or "").strip()
        if not text:
            return {}
        delimiter = ";" if text.count(";") > text.count(",") else ","
        values = next(csv.reader([text], delimiter=delimiter), [])
        return {f"Columna {index}": item for index, item in enumerate(values, start=1)}

    @staticmethod
    def _first_value(rows: list[dict], key: str) -> str:
        for row in rows:
            value = row.get(key)
            if value not in (None, ""):
                return str(value)
        return ""

    @staticmethod
    def _slugify(value: str) -> str:
        normalized = normalize_text(value).replace(" ", "-")
        clean = "".join(char for char in normalized if char.isalnum() or char in {"-", "_"})
        return clean.strip("-_") or "export"

    def _build_power_bi_workbook(
        self,
        *,
        table_rows: dict[str, list[dict]],
        relationships: list[dict],
        export_id: str,
        exported_at: str,
        garage_id: str,
        generated_by: dict,
    ) -> io.BytesIO:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        garage = self._garage_info(garage_id)
        generated_at = parse_datetime(exported_at) or utcnow()
        logo_path = self._default_logo_path()

        sheet_order = ["Resumen", "Historico", "Ocupacion", "Ingresos"]
        for sheet_name in sheet_order:
            worksheet = workbook.create_sheet(title=sheet_name)
            self._populate_sheet(
                worksheet,
                table_rows.get(sheet_name, []),
                title=f"Reporte de Parkings - {sheet_name}",
                garage=garage,
                generated_by=generated_by,
                generated_at=generated_at,
                logo_path=logo_path,
            )

        meta_sheet = workbook.create_sheet(title="Meta")
        metadata_rows = [
            {"key": "export_id", "value": export_id},
            {"key": "exported_at", "value": exported_at},
            {"key": "garage_id", "value": garage_id},
            {"key": "requested_format", "value": "pbix"},
            {"key": "actual_format", "value": "power-bi-import-bundle"},
        ]
        self._populate_sheet(
            meta_sheet,
            metadata_rows,
            title="Metadatos del reporte",
            garage=garage,
            generated_by=generated_by,
            generated_at=generated_at,
            logo_path=logo_path,
        )

        relationships_sheet = workbook.create_sheet(title="Relaciones")
        self._populate_sheet(
            relationships_sheet,
            self._coerce_rows(relationships),
            title="Relaciones sugeridas",
            garage=garage,
            generated_by=generated_by,
            generated_at=generated_at,
            logo_path=logo_path,
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def _populate_sheet(
        self,
        worksheet,
        rows: list[dict],
        *,
        title: str,
        garage: dict,
        generated_by: dict,
        generated_at: datetime,
        logo_path: Path | None,
    ) -> None:
        render_report_sheet(
            worksheet,
            title=title,
            rows=rows,
            garage=garage,
            generated_by=generated_by,
            generated_at=generated_at,
            table_name=f"SmartPark_{worksheet.title}",
            logo_path=logo_path,
        )

    @staticmethod
    def _excel_value(value):
        if value is None:
            return ""
        if isinstance(value, (int, float, bool)):
            return value
        if isinstance(value, datetime):
            return value.isoformat()
        return str(value)

    @staticmethod
    def _garage_info(garage_id: str) -> dict:
        filters = [{"column": "id", "value": garage_id, "optional": True}]
        for table_name in ("garages", "garajes"):
            try:
                rows = select_rows(table_name, filters=filters, limit=1)
                if not rows:
                    rows = select_rows(
                        table_name,
                        filters=[{"column": "garage_id", "value": garage_id, "optional": True}],
                        limit=1,
                    )
            except Exception:
                rows = []
            if rows:
                row = rows[0]
                name = row.get("company_name") or row.get("nombre") or row.get("name") or "SmartPark Garage"
                address = row.get("address") or row.get("direccion") or row.get("company_address") or ""
                return {**row, "name": name, "nombre": name, "address": address, "direccion": address}
        return {"id": garage_id, "garage_id": garage_id, "name": "SmartPark Garage", "address": ""}

    @staticmethod
    def _default_logo_path() -> Path | None:
        path = Path(__file__).resolve().parents[2] / "frontend" / "public" / "images" / "logo-smartpark.png"
        return path if path.exists() else None

    @staticmethod
    def _build_power_bi_readme(*, period_label: str, floor_label: str) -> str:
        return (
            "SmartPark Power BI Import Bundle\n"
            "================================\n\n"
            "Este paquete se genera como alternativa a un .pbix real.\n"
            "Power BI Desktop no ofrece una ruta oficial soportada para crear o convertir .pbix por código en backend.\n\n"
            f"Periodo: {period_label}\n"
            f"Filtro de piso: {floor_label}\n\n"
            "Contenido:\n"
            "- smartpark-powerbi.xlsx: workbook con sheets Resumen, Historico, Ocupacion, Ingresos, Meta y Relaciones.\n"
            "- smartpark-powerbi.json: mismas tablas en JSON.\n"
            "- relationships.json: relaciones sugeridas para el modelo.\n"
            "- manifest.json: metadatos del bundle.\n\n"
            "Uso recomendado en Power BI Desktop:\n"
            "1. Importar smartpark-powerbi.xlsx o smartpark-powerbi.json.\n"
            "2. Crear relaciones usando export_id y floor_key segun relationships.json.\n"
            "3. Guardar el proyecto en Power BI Desktop como .pbix.\n"
        )
