from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from utils.supabase_client import parse_datetime, utcnow


HEADER_ROW = 4
DATA_START_ROW = HEADER_ROW + 1
LEGAL_FOOTER = "Documento generado por SmartPark. Uso interno sujeto a validacion administrativa."

BRAND_FILL = "0F172A"
ACCENT_FILL = "06B6D4"
HIGHLIGHT_FILL = "F59E0B"
HEADER_FILL = "0F172A"
ALT_ROW_FILL = "F8FAFC"
META_FILL = "FFFFFF"
FOOTER_FILL = "EAF8F2"
TEXT_DARK = "0F172A"
TEXT_MUTED = "475569"
WHITE = "FFFFFF"
BORDER_COLOR = "D7DEE8"

DATE_HEADERS = {
    "fecha",
    "date",
    "created_at",
    "updated_at",
    "exported_at",
    "entrada",
    "salida",
    "entry",
    "exit",
}
CURRENCY_TOKENS = ("monto", "amount", "income", "ingreso", "revenue", "dop")


def render_report_sheet(
    worksheet,
    *,
    title: str,
    rows: list[dict[str, Any]],
    garage: dict[str, Any] | None = None,
    generated_by: dict[str, Any] | None = None,
    generated_at: datetime | None = None,
    table_name: str | None = None,
    logo_path: str | Path | None = None,
) -> None:
    generated_at = generated_at or utcnow()
    garage = garage or {}
    generated_by = generated_by or {}

    headers = _headers_from_rows(rows)
    max_column = max(len(headers), 4)

    _setup_page(worksheet)
    _render_header(
        worksheet,
        title=title,
        max_column=max_column,
        garage=garage,
        generated_by=generated_by,
        generated_at=generated_at,
    )

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=HEADER_ROW, column=column_index, value=header)
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    for row_index, row in enumerate(rows, start=DATA_START_ROW):
        if row_index % 2 == 0:
            for column_index in range(1, len(headers) + 1):
                worksheet.cell(row=row_index, column=column_index).fill = PatternFill(fill_type="solid", fgColor=ALT_ROW_FILL)

        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=_excel_value(row.get(header), header))
            _format_data_cell(cell, header)

    if not rows:
        worksheet.cell(row=DATA_START_ROW, column=1, value="Sin datos")
        worksheet.cell(row=DATA_START_ROW, column=1).font = Font(italic=True, color=TEXT_MUTED)
        worksheet.cell(row=DATA_START_ROW, column=1).border = _thin_border()

    footer_row = DATA_START_ROW + max(len(rows), 1) + 1
    _render_footer(worksheet, footer_row=footer_row, headers=headers, record_count=len(rows))

    last_data_row = DATA_START_ROW + max(len(rows), 1) - 1
    _add_excel_table(
        worksheet,
        table_name=table_name or title,
        min_col=1,
        max_col=len(headers),
        header_row=HEADER_ROW,
        last_data_row=last_data_row,
    )
    _autosize_columns(worksheet, headers=headers, rows=rows, max_column=max_column)
    worksheet.freeze_panes = f"A{DATA_START_ROW}"
    worksheet.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{last_data_row}"
    worksheet.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"


def _setup_page(worksheet) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_margins.left = 0.35
    worksheet.page_margins.right = 0.35
    worksheet.page_margins.top = 0.55
    worksheet.page_margins.bottom = 0.55


def _render_header(
    worksheet,
    *,
    title: str,
    max_column: int,
    garage: dict[str, Any],
    generated_by: dict[str, Any],
    generated_at: datetime,
) -> None:
    title_end_column = max(1, max_column - 2)
    date_start_column = max(2, max_column - 1)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_end_column)
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, color=TEXT_DARK, size=20)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    if date_start_column <= max_column:
        worksheet.merge_cells(start_row=1, start_column=date_start_column, end_row=1, end_column=max_column)
        generated_cell = worksheet.cell(row=1, column=date_start_column, value=f"{generated_at:%d/%m/%Y %H:%M}")
        generated_cell.font = Font(bold=True, color=TEXT_MUTED, size=10)
        generated_cell.alignment = Alignment(horizontal="right", vertical="center")

    garage_name = _first(garage, "name", "nombre", "company_name", default="SmartPark Garage")
    garage_address = _first(garage, "address", "direccion", "company_address", default="Direccion no registrada")
    user_name = _first(generated_by, "name", "nombre", "email", default="Usuario no disponible")
    user_email = _first(generated_by, "email", default="")

    meta_values = [
        ("Garage", garage_name),
        ("Direccion", garage_address),
        ("Usuario", f"{user_name} ({user_email})" if user_email and user_email != user_name else user_name),
    ]
    _render_meta_band(worksheet, max_column=max_column, items=meta_values)

    worksheet.row_dimensions[1].height = 28
    worksheet.row_dimensions[2].height = 20
    worksheet.row_dimensions[3].height = 8
    worksheet.row_dimensions[HEADER_ROW].height = 24


def _try_add_logo(worksheet, logo_path: str | Path | None) -> None:
    if not logo_path:
        return
    path = Path(logo_path)
    if not path.exists():
        return
    try:
        from openpyxl.drawing.image import Image

        image = Image(str(path))
        image.height = 30
        image.width = 106
        worksheet.add_image(image, "A1")
        worksheet.cell(row=1, column=1).value = None
    except Exception:
        return


def _render_meta_band(worksheet, *, max_column: int, items: list[tuple[str, str]]) -> None:
    for column_index in range(1, max_column + 1):
        cell = worksheet.cell(row=2, column=column_index)
        cell.fill = PatternFill(fill_type="solid", fgColor=META_FILL)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    if max_column >= 6:
        pairs = [(1, 2), (3, 4), (5, 6)]
        for (label, value), (label_col, value_col) in zip(items, pairs):
            label_cell = worksheet.cell(row=2, column=label_col, value=label)
            value_cell = worksheet.cell(row=2, column=value_col, value=value)
            label_cell.font = Font(bold=True, color=TEXT_DARK, size=9)
            value_cell.font = Font(color=TEXT_MUTED, size=9)
            value_cell.alignment = Alignment(horizontal="left", vertical="center")
        if max_column > 6:
            worksheet.merge_cells(start_row=2, start_column=6, end_row=2, end_column=max_column)
        return

    metadata = " | ".join(f"{label}: {value}" for label, value in items)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)
    cell = worksheet.cell(row=2, column=1, value=metadata)
    cell.font = Font(color=TEXT_MUTED, size=9)


def _render_footer(worksheet, *, footer_row: int, headers: list[str], record_count: int) -> None:
    max_column = max(len(headers), 4)
    for column_index in range(1, max_column + 1):
        cell = worksheet.cell(row=footer_row, column=column_index)
        cell.fill = PatternFill(fill_type="solid", fgColor=FOOTER_FILL)
        cell.border = _thin_border()

    worksheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=min(2, max_column))
    count_cell = worksheet.cell(row=footer_row, column=1, value=f"Total de registros: {record_count}")
    count_cell.font = Font(bold=True, color=TEXT_DARK)
    count_cell.alignment = Alignment(horizontal="left")


def _add_excel_table(worksheet, *, table_name: str, min_col: int, max_col: int, header_row: int, last_data_row: int) -> None:
    safe_name = _safe_table_name(table_name)
    ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{last_data_row}"
    table = Table(displayName=safe_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _autosize_columns(worksheet, *, headers: list[str], rows: list[dict[str, Any]], max_column: int) -> None:
    for column_index in range(1, max_column + 1):
        header = headers[column_index - 1] if column_index <= len(headers) else ""
        values = [str(header)]
        for row in rows:
            values.append("" if row.get(header) is None else str(row.get(header)))
        width = min(max(max((len(item) for item in values), default=10) + 2, 12), 42)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _headers_from_rows(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return ["info"]
    headers: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(str(key))
    return headers


def _format_data_cell(cell, header: str) -> None:
    cell.border = _thin_border()
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.font = Font(color=TEXT_DARK, size=10)
    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if _is_currency_header(header):
            cell.number_format = '"RD$" #,##0.00'
        else:
            cell.number_format = '#,##0' if isinstance(cell.value, int) else '#,##0.00'
    elif isinstance(cell.value, datetime):
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = "dd/mm/yyyy"


def _excel_value(value: Any, header: str) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value) if _is_currency_header(header) else value

    parsed = parse_datetime(value)
    if parsed and _is_date_header(header):
        return parsed.replace(tzinfo=None)

    if _is_currency_header(header):
        number = _parse_number(value)
        if number is not None:
            return number

    return str(value)


def _parse_number(value: Any) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    clean = text.replace("RD$", "").replace("$", "").replace(",", "").strip()
    try:
        return float(clean)
    except ValueError:
        return None


def _is_date_header(header: str) -> bool:
    normalized = str(header or "").strip().lower()
    return normalized in DATE_HEADERS or normalized.endswith("_at") or "fecha" in normalized


def _is_currency_header(header: str) -> bool:
    normalized = str(header or "").strip().lower()
    return any(token in normalized for token in CURRENCY_TOKENS)


def _safe_table_name(value: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_]", "_", str(value or "SmartParkTable"))
    name = re.sub(r"_+", "_", name).strip("_") or "SmartParkTable"
    if name[0].isdigit():
        name = f"T_{name}"
    return name[:240]


def _first(source: dict[str, Any], *keys: str, default: str = "") -> str:
    for key in keys:
        value = source.get(key)
        if value not in (None, ""):
            return str(value)
    return default


def _thin_border() -> Border:
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)
